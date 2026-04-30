"""Excel parsing tool for weekly report pipeline."""

from __future__ import annotations

import base64
import binascii
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any

from google.adk.tools.tool_context import ToolContext
from pydantic import ValidationError

from ._utils import _format_validation_error
from ..schemas.model import TaskItem

_SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
_STATUS_VALUES = {"완료", "진행", "진행중", "진행 중", "예정", "지연"}
_ISSUE_KEYWORDS = ["이슈", "리스크", "지연", "문제", "오류", "블로킹", "장애", "실패", "확정 안됨"]
_SHEET_SCAN_ROWS = 250
_YEAR_2000_BASE = 2000
_OPENPYXL_INDEXED_RED = 10
_OPENPYXL_INDEXED_BLUE = 12
_NAME_STOPWORDS = {
    "업무",
    "전체",
    "관리",
    "수행",
    "센터",
    "본사",
    "지원",
    "상시",
    "대응",
    "진행",
    "완료",
    "예정",
    "프로젝트",
    "구축",
}


def _extract_inline_bytes(raw_data: Any) -> bytes | None:
    if raw_data is None:
        return None
    if isinstance(raw_data, bytes):
        return raw_data
    if isinstance(raw_data, str):
        try:
            return base64.b64decode(raw_data, validate=True)
        except (binascii.Error, ValueError):
            try:
                return base64.urlsafe_b64decode(raw_data)
            except (binascii.Error, ValueError):
                return raw_data.encode("utf-8")
    return None


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_member_name(text: str) -> bool:
    # 일반적으로 한국인 이름은 2~3자이며, 4자는 업무 용어 오탐이 많아 제외한다.
    return bool(re.fullmatch(r"[가-힣]{2,3}", text))


def _extract_member_names(*texts: str) -> list[str]:
    names: list[str] = []
    for text in texts:
        if not text:
            continue
        for group in re.findall(r"\(([가-힣,\s]+)\)", text):
            for token in re.split(r"[,\s]+", group):
                token = token.strip()
                if _is_member_name(token):
                    names.append(token)
        for token in re.findall(r"[가-힣]{2,4}", text):
            if _is_member_name(token) and token not in _NAME_STOPWORDS:
                names.append(token)
    deduped: list[str] = []
    for name in names:
        if name not in deduped:
            deduped.append(name)
    return deduped


def _infer_center_from_filename(path: Path) -> str:
    m = re.search(r"([가-힣A-Za-z0-9]+센터)", path.stem)
    if m:
        return m.group(1)
    return ""


def _extract_center_from_text(text: str) -> str:
    if not text:
        return ""
    m = re.search(r"([가-힣A-Za-z0-9]+센터)", text)
    if m:
        return m.group(1)
    return ""


def _infer_center_from_sheet(ws: Any) -> str:
    """
    파일명으로 센터를 못 찾은 경우 시트 본문에서 센터명을 추론한다.
    우선순위: '<센터> 개인업무' 행 -> 임의의 '<...센터>' 텍스트
    """
    row_cells_all = list(ws.iter_rows())
    for row_cells in row_cells_all[:_SHEET_SCAN_ROWS]:
        cells = [_normalize_cell(c.value) for c in row_cells]
        joined = " ".join(c for c in cells if c)
        if "개인업무" in joined and "센터" in joined:
            center = _extract_center_from_text(joined)
            if center:
                return center

    for row_cells in row_cells_all[:_SHEET_SCAN_ROWS]:
        cells = [_normalize_cell(c.value) for c in row_cells]
        joined = " ".join(c for c in cells if c)
        center = _extract_center_from_text(joined)
        if center:
            return center

    return ""


def _parse_short_date(value: Any) -> str:
    text = _normalize_cell(value)
    if not text:
        return ""
    if text in {"-", ".", "상시대응", "상시 대응", "미정"}:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    raw_text = text
    if re.match(r"\d{4}-\d{2}-\d{2}", raw_text):
        return raw_text[:10]
    if re.match(r"\d{4}\.\d{2}\.\d{2}", raw_text):
        yyyy, mm, dd = raw_text[:10].split(".")
        return f"{int(yyyy):04d}-{int(mm):02d}-{int(dd):02d}"

    text = text.replace("/", ".").replace("-", ".")
    if re.fullmatch(r"\d{2}\.\d{2}\.\d{2}", text):
        yy, mm, dd = text.split(".")
        return f"{_YEAR_2000_BASE + int(yy):04d}-{int(mm):02d}-{int(dd):02d}"
    if re.fullmatch(r"\d{4}\.\d{2}\.\d{2}", text):
        yyyy, mm, dd = text.split(".")
        return f"{int(yyyy):04d}-{int(mm):02d}-{int(dd):02d}"
    return ""


def _normalize_rgb_to_hex6(raw_rgb: Any) -> str:
    if not isinstance(raw_rgb, str):
        return ""
    value = raw_rgb.strip().upper()
    if len(value) == 8:
        value = value[-6:]
    if len(value) != 6 or any(c not in "0123456789ABCDEF" for c in value):
        return ""
    return value


def _is_red_hex(hex6: str) -> bool:
    if not hex6:
        return False
    try:
        r = int(hex6[0:2], 16)
        g = int(hex6[2:4], 16)
        b = int(hex6[4:6], 16)
    except ValueError:
        return False
    return r >= 160 and (r - g) >= 40 and (r - b) >= 40


def _is_blue_hex(hex6: str) -> bool:
    if not hex6:
        return False
    try:
        r = int(hex6[0:2], 16)
        g = int(hex6[2:4], 16)
        b = int(hex6[4:6], 16)
    except ValueError:
        return False
    return b >= 140 and (b - r) >= 30 and (b - g) >= 10


def _extract_cell_colors(cell: Any) -> list[Any]:
    """셀의 fill/font에서 색상 객체 목록을 추출한다."""
    colors: list[Any] = []
    fill = getattr(cell, "fill", None)
    if fill is not None:
        colors.extend([
            getattr(fill, "fgColor", None),
            getattr(fill, "bgColor", None),
            getattr(fill, "start_color", None),
            getattr(fill, "end_color", None),
        ])
    font = getattr(cell, "font", None)
    if font is not None:
        colors.append(getattr(font, "color", None))
    return colors


def _cell_has_color_style(cell: Any, hex_checker, indexed_value: int) -> bool:
    """셀에 특정 색상(hex_checker 기준 또는 indexed_value)이 있는지 판별한다."""
    if cell is None:
        return False
    for color in _extract_cell_colors(cell):
        if color is None:
            continue
        hex6 = _normalize_rgb_to_hex6(getattr(color, "rgb", None))
        if hex6 and hex_checker(hex6):
            return True
        if getattr(color, "indexed", None) == indexed_value:
            return True
    return False


def _cell_has_red_style(cell: Any) -> bool:
    """셀의 fill/font 색상 중 빨강 계열 존재 여부를 판별한다."""
    return _cell_has_color_style(cell, _is_red_hex, _OPENPYXL_INDEXED_RED)


def _cell_has_blue_style(cell: Any) -> bool:
    """셀의 fill/font 색상 중 파랑 계열 존재 여부를 판별한다."""
    return _cell_has_color_style(cell, _is_blue_hex, _OPENPYXL_INDEXED_BLUE)


def _color_obj_has_visible_color(color: Any) -> bool:
    if color is None:
        return False
    hex6 = _normalize_rgb_to_hex6(getattr(color, "rgb", None))
    if hex6 and hex6 not in {"000000", "FFFFFF"}:
        return True
    indexed = getattr(color, "indexed", None)
    if indexed is not None and indexed not in {0, 1, 64}:
        return True
    theme = getattr(color, "theme", None)
    tint = getattr(color, "tint", None)
    if theme is not None and tint not in (None, 0):
        return True
    return False


def _cell_has_any_color_style(cell: Any) -> bool:
    """셀에 기본값이 아닌 색상 스타일이 있는지 판별한다."""
    if cell is None:
        return False

    fill = getattr(cell, "fill", None)
    if fill is not None:
        if getattr(fill, "patternType", None) not in (None, "none"):
            for attr in ("fgColor", "bgColor", "start_color", "end_color"):
                if _color_obj_has_visible_color(getattr(fill, attr, None)):
                    return True

    font = getattr(cell, "font", None)
    if font is not None and _color_obj_has_visible_color(getattr(font, "color", None)):
        return True

    return False


def _row_has_red_style(row_cells: list[Any], candidate_indexes: list[int]) -> bool:
    for idx in candidate_indexes:
        if idx < 0 or idx >= len(row_cells):
            continue
        if _cell_has_red_style(row_cells[idx]):
            return True
    return False


def _extract_member_name_from_header_text(text: str) -> str:
    """개인업무 블록 헤더 텍스트에서 담당자명을 추출한다."""
    if not text:
        return ""
    head = text.split("-", 1)[0].strip()
    if _is_member_name(head):
        return head
    names = _extract_member_names(text)
    return names[0] if names else ""


def _is_member_header_fallback(project_cell: str, summary_cell: str, status_cell: str, member_candidate: str) -> bool:
    """색상 인식 실패 시 텍스트 패턴으로 개인업무 헤더를 보완 인식한다."""
    if not member_candidate or status_cell:
        return False
    source = project_cell or summary_cell
    if not source:
        return False
    if "전체 업무" in source or "개인업무" in source:
        return True
    return False


def _extract_member_header_from_row(row_cells: list[Any], status_cell: str) -> tuple[str, bool]:
    """
    행 전체를 스캔해 개인업무 헤더(이름) 후보를 찾는다.
    반환: (member_name, is_colored_header)
    """
    if _normalize_status_text(status_cell):
        if _is_valid_status_text(status_cell):
            return "", False

    colored_candidate = ""
    plain_candidate = ""
    joined_row = " ".join(_normalize_cell(getattr(c, "value", None)) for c in row_cells)
    has_level1_marker = any(_normalize_cell(getattr(c, "value", None)).strip().startswith("-") for c in row_cells)

    # 이름 헤더는 보통 좌측 컬럼에 오므로 앞 3개 셀까지만 본다.
    for cell in row_cells[:3]:
        text = _normalize_cell(getattr(cell, "value", None))
        if not text:
            continue
        stripped = text.strip()
        if stripped.startswith((">", "*", ":")):
            continue
        candidate = _extract_member_name_from_header_text(text)
        if not candidate:
            continue
        # 이름 헤더는 파란색으로 강조된다는 작성 규칙을 우선 적용한다.
        if _cell_has_blue_style(cell):
            # 담당자 헤더는 아래 패턴일 때 허용:
            # 1) 전체 업무/개인업무 문구
            # 2) 같은 행에 Level1('- ...') 프로젝트 헤더가 존재
            if "전체 업무" in joined_row or "개인업무" in joined_row or has_level1_marker:
                colored_candidate = candidate
                break
        if not plain_candidate and ("전체 업무" in joined_row or "개인업무" in joined_row or has_level1_marker):
            plain_candidate = candidate

    if colored_candidate:
        return colored_candidate, True
    if plain_candidate:
        return plain_candidate, False
    return "", False


def _normalize_status_text(value: str) -> str:
    text = value.strip()
    text = re.sub(r"[\[\]\(\)]", "", text)
    text = text.replace(" ", "")
    if text.startswith("진행"):
        return "진행"
    if text.startswith("완료"):
        return "완료"
    if text.startswith("예정"):
        return "예정"
    if text.startswith("지연"):
        return "지연"
    return value.strip()


def _is_valid_status_text(value: str) -> bool:
    return _normalize_status_text(value) in _STATUS_VALUES


def _extract_status_from_text(*texts: str) -> str:
    for text in texts:
        if not text:
            continue
        m = re.search(r"\[(완료|진행|예정|지연)\]", text)
        if m:
            return m.group(1)
        normalized = _normalize_status_text(text)
        if _is_valid_status_text(normalized):
            return normalized
    return ""


def _parse_month_day_token(token: str) -> str:
    token = token.strip().replace(".", "/").replace("-", "/")
    if re.fullmatch(r"\d{1,2}/\d{1,2}", token):
        mm, dd = token.split("/")
        year = date.today().year
        return f"{year:04d}-{int(mm):02d}-{int(dd):02d}"
    return _parse_short_date(token)


def _extract_date_range_from_text(*texts: str) -> tuple[str, str]:
    pattern = r"(\d{1,2}[./-]\d{1,2}|\d{2}\.\d{2}\.\d{2}|\d{4}[./-]\d{1,2}[./-]\d{1,2})\s*~\s*(\d{1,2}[./-]\d{1,2}|\d{2}\.\d{2}\.\d{2}|\d{4}[./-]\d{1,2}[./-]\d{1,2})"
    for text in texts:
        if not text:
            continue
        m = re.search(pattern, text)
        if not m:
            continue
        return _parse_month_day_token(m.group(1)), _parse_month_day_token(m.group(2))
    return "", ""


def _extract_task_text_from_row(cells: list[str], project_cell: str, summary_cell: str) -> str:
    candidates = [summary_cell, project_cell]
    candidates.extend(cells)
    for text in candidates:
        if not text:
            continue
        normalized = text.strip()
        if normalized.startswith((">", "*", ":", "-")):
            return normalized
    for text in candidates:
        if text and len(text.strip()) >= 2:
            return text.strip()
    return ""


def _strip_bullet_prefix(text: str) -> str:
    return re.sub(r"^\s*[-*>:]+\s*", "", text or "").strip()


def _extract_level1_project_from_row(cells: list[str]) -> str:
    """
    Level 1(-) 프로젝트명을 행 전체에서 탐지한다.
    개인업무 담당자 헤더(- 전체 업무 관리)는 제외한다.
    """
    for cell_text in cells:
        t = (cell_text or "").strip()
        if not t.startswith("-"):
            continue
        normalized = _strip_bullet_prefix(t)
        if not normalized:
            continue
        if "전체 업무 관리" in normalized or normalized == "전체 업무 관리":
            continue
        return normalized
    return ""


def _infer_hierarchical_columns(row_cells_all: list[Any]) -> tuple[int, int, int]:
    """헤더가 없을 때 status/summary/project 컬럼을 데이터 분포로 추정한다."""
    if not row_cells_all:
        return 0, 1, 2

    max_cols = max(len(r) for r in row_cells_all)
    status_scores = [0] * max_cols

    for row_cells in row_cells_all[:150]:
        for idx, cell in enumerate(row_cells):
            cell_text = _normalize_cell(getattr(cell, "value", None))
            if _is_valid_status_text(cell_text):
                status_scores[idx] += 1

    status_col = 1
    best_score = max(status_scores) if status_scores else 0
    if best_score > 0:
        status_col = status_scores.index(best_score)

    summary_scores = [0] * max_cols
    for row_cells in row_cells_all[:200]:
        status_text = _normalize_cell(row_cells[status_col].value) if status_col < len(row_cells) else ""
        if not _is_valid_status_text(status_text):
            continue
        for idx, cell in enumerate(row_cells):
            if idx == status_col:
                continue
            text = _normalize_cell(getattr(cell, "value", None))
            if len(text) >= 4:
                summary_scores[idx] += len(text)

    summary_col = min(status_col + 1, max_cols - 1) if max_cols > 0 else 2
    if summary_scores and max(summary_scores) > 0:
        summary_col = summary_scores.index(max(summary_scores))

    project_col = max(0, status_col - 1)
    return project_col, status_col, summary_col


async def _resolve_excel_path_from_artifact(tool_context: ToolContext) -> tuple[str | None, str | None]:
    try:
        artifact_names = await tool_context.list_artifacts()
    except Exception:
        return None, None

    excel_candidates = [
        name
        for name in artifact_names
        if Path(name.split(":", 1)[-1]).suffix.lower() in _SUPPORTED_EXCEL_EXTENSIONS
    ]
    if not excel_candidates:
        return None, None

    for artifact_name in reversed(excel_candidates):
        artifact = await tool_context.load_artifact(artifact_name)
        if artifact is None and not artifact_name.startswith("user:"):
            artifact = await tool_context.load_artifact(f"user:{artifact_name}")
        if artifact is None or artifact.inline_data is None:
            continue

        data = _extract_inline_bytes(artifact.inline_data.data)
        if not data:
            continue

        source_name = Path(artifact_name.split(":", 1)[-1]).name
        upload_dir = Path("uploads")
        upload_dir.mkdir(parents=True, exist_ok=True)
        temp_path = upload_dir / f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{source_name}"
        temp_path.write_bytes(data)
        return str(temp_path), artifact_name

    return None, None


def _build_column_map(headers: list[str]) -> dict[str, int]:
    aliases: dict[str, list[str]] = {
        "project_name": ["프로젝트명", "프로젝트", "project", "project_name", "업무구분"],
        "center_name": ["센터", "센터명", "center", "center_name", "소속센터", "본부"],
        "member_name": ["담당자", "작성자", "이름", "성명", "member", "member_name", "사원명"],
        "status": ["상태", "status", "진행상태", "진행 상태"],
        "summary": ["업무내용", "내용", "summary", "업무 내용", "작업내용", "업무명", "요약"],
        "prev_start": ["이전시작", "시작일", "prev_start", "이전 시작", "착수일", "전주실적 시작일"],
        "prev_end": ["이전종료", "종료일", "prev_end", "이전 종료", "완료일", "전주실적 종료일"],
        "next_start": ["다음시작", "next_start", "다음 시작", "차주시작", "금주계획 시작일"],
        "next_end": ["다음종료", "next_end", "다음 종료", "차주종료", "금주계획 종료일"],
    }
    col_map: dict[str, int] = {}
    for field, alias_list in aliases.items():
        for idx, header in enumerate(headers):
            if header in alias_list:
                col_map[field] = idx
                break
    return col_map


def _parse_row(
    row_data: dict,
    col_map: dict[str, int],
    inferred_center: str,
    style_delayed: bool = False,
) -> dict | None:
    headers = list(row_data.keys())

    def get(field: str) -> Any:
        idx = col_map.get(field)
        if idx is not None and idx < len(headers):
            return row_data.get(headers[idx])
        return None

    project_name = str(get("project_name") or "").strip()
    status = str(get("status") or "").strip()
    summary = str(get("summary") or "").strip()
    if not project_name or not status or not summary:
        return None

    prev_end_val = get("prev_end")
    is_delayed = False
    if status == "지연" or style_delayed:
        is_delayed = True
    elif status in ("진행", "진행중", "진행 중") and prev_end_val:
        parsed_prev_end = _parse_short_date(prev_end_val)
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", parsed_prev_end):
            try:
                end_date = datetime.strptime(parsed_prev_end, "%Y-%m-%d").date()
                is_delayed = end_date < date.today()
            except Exception:
                pass

    has_issue_keyword = any(kw in summary for kw in _ISSUE_KEYWORDS)
    member_name = str(get("member_name") or "").strip()
    if not member_name:
        members = _extract_member_names(summary, project_name)
        member_name = ", ".join(members)
    center_name = str(get("center_name") or "").strip() or inferred_center
    if not member_name or not center_name:
        return None

    return {
        "project_name": project_name,
        "center_name": center_name,
        "member_name": member_name,
        "status": status,
        "summary": summary,
        "prev_start": _parse_short_date(get("prev_start")),
        "prev_end": _parse_short_date(get("prev_end")),
        "next_start": _parse_short_date(get("next_start")),
        "next_end": _parse_short_date(get("next_end")),
        "is_delayed": is_delayed,
        "has_issue_keyword": has_issue_keyword,
    }


def _collect_header_identity_map(
    ws: Any,
    headers: list[str],
    col_map: dict[str, int],
) -> tuple[dict[int, dict[str, str]], list[dict]]:
    """
    헤더 기반으로 row별 member만 수집한다.
    center는 파일명에서 추론한 값을 사용한다.
    """
    identity_map: dict[int, dict[str, str]] = {}
    anomalies: list[dict] = []

    member_idx = col_map.get("member_name")

    # member 컬럼이 없는 형식이면 헤더 기반 member 보정은 수행하지 않는다.
    if member_idx is None:
        return identity_map, anomalies

    for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
        row_values = [c.value for c in row_cells]
        if all(v is None for v in row_values):
            continue
        row_data = dict(zip(headers, row_values))

        member_name = ""
        if member_idx is not None and member_idx < len(headers):
            member_name = str(row_data.get(headers[member_idx]) or "").strip()

        payload: dict[str, str] = {}
        if member_name:
            payload["member_name"] = member_name

        if payload:
            identity_map[row_idx] = payload

    return identity_map, anomalies


def _parse_hierarchical_rows(ws: Any, inferred_center: str) -> tuple[list[dict], list[dict]]:
    row_cells_all = list(ws.iter_rows())
    if not row_cells_all:
        return [], []

    header_idx = None
    col_project, col_status, col_summary = 0, 1, 2
    col_prev_start, col_prev_end, col_next_start, col_next_end = 3, 4, 5, 6

    for idx, row_cells in enumerate(row_cells_all):
        cells = [_normalize_cell(c.value) for c in row_cells]
        if "업무구분" in cells and "상태" in cells and "요약" in cells:
            header_idx = idx
            col_project = cells.index("업무구분")
            col_status = cells.index("상태")
            col_summary = cells.index("요약")
            col_prev_start = col_summary + 1
            col_prev_end = col_summary + 2
            col_next_start = col_summary + 3
            col_next_end = col_summary + 4
            break

    if header_idx is None:
        col_project, col_status, col_summary = _infer_hierarchical_columns(row_cells_all)
        col_prev_start = col_summary + 1
        col_prev_end = col_summary + 2
        col_next_start = col_summary + 3
        col_next_end = col_summary + 4

    start_idx = (header_idx + 1) if header_idx is not None else 1
    initial_member = ""
    # "<센터> 개인업무" 마커 다음 행부터 개인업무 영역을 읽는다.
    for idx, row_cells in enumerate(row_cells_all):
        cells = [_normalize_cell(c.value) for c in row_cells]
        joined_row = " ".join(c for c in cells if c)
        if "개인업무" in joined_row and "센터" in joined_row:
            # 마커 행에 담당자명이 함께 기재되는 경우(예: 이용필) 초기 담당자로 설정한다.
            row_member_candidate, _ = _extract_member_header_from_row(list(row_cells), "")
            if row_member_candidate:
                initial_member = row_member_candidate
            start_idx = max(start_idx, idx + 1)
            break

    records: list[dict] = []
    anomalies: list[dict] = []
    current_project = ""
    current_center = inferred_center
    current_member = initial_member
    known_members: set[str] = {initial_member} if initial_member else set()
    last_record: dict[str, Any] | None = None

    for row_idx, row_cells in enumerate(row_cells_all[start_idx:], start=start_idx + 1):
        cells = [_normalize_cell(c.value) for c in row_cells]
        if not any(cells):
            continue

        level1_project = _extract_level1_project_from_row(cells)
        if level1_project:
            current_project = level1_project
            # Level1 프로젝트 행에 담당자(파란색 이름)가 함께 있으면 멤버 컨텍스트도 갱신한다.
            row_member_candidate, row_member_colored = _extract_member_header_from_row(list(row_cells), status_cell="")
            if row_member_candidate and row_member_colored:
                current_member = row_member_candidate
                known_members.add(current_member)
            continue

        project_cell = cells[col_project] if col_project < len(cells) else ""
        status_cell = cells[col_status] if col_status < len(cells) else ""
        summary_cell = cells[col_summary] if col_summary < len(cells) else ""
        prev_start_cell = cells[col_prev_start] if col_prev_start < len(cells) else ""
        prev_end_cell = cells[col_prev_end] if col_prev_end < len(cells) else ""
        next_start_cell = cells[col_next_start] if col_next_start < len(cells) else ""
        next_end_cell = cells[col_next_end] if col_next_end < len(cells) else ""
        joined = " ".join(c for c in (project_cell, status_cell, summary_cell) if c)
        row_text = " ".join(c for c in cells if c)

        if "센터" in joined and "개인업무" in joined:
            current_center = _extract_center_from_text(joined) or project_cell or joined or current_center
            continue
        if project_cell and ("센터" in project_cell or "본부" in project_cell) and not status_cell:
            current_center = _extract_center_from_text(project_cell) or project_cell
            continue

        if project_cell and not status_cell and not summary_cell:
            if "프로젝트" in project_cell or "구축" in project_cell or "유지보수" in project_cell or project_cell in {"기타", "학습"}:
                current_project = project_cell
                continue

        # 이름은 한 번만 나오고 이후 행이 이어지는 형태가 많아, 행 전체를 스캔해 헤더를 우선 인식한다.
        row_member_candidate, row_member_colored = _extract_member_header_from_row(list(row_cells), status_cell)
        member_candidate = (
            row_member_candidate
            or _extract_member_name_from_header_text(project_cell)
            or _extract_member_name_from_header_text(summary_cell)
        )
        is_colored_member_header = bool(row_member_colored and member_candidate)
        is_member_header_by_pattern = _is_member_header_fallback(
            project_cell=project_cell,
            summary_cell=summary_cell,
            status_cell=status_cell,
            member_candidate=member_candidate,
        )
        if is_colored_member_header or is_member_header_by_pattern:
            current_member = member_candidate
            if current_member:
                known_members.add(current_member)
            continue

        inferred_status = _extract_status_from_text(status_cell, summary_cell, project_cell, row_text)
        status_cell = _normalize_status_text(inferred_status or status_cell)
        task_text = _extract_task_text_from_row(cells, project_cell, summary_cell)

        # Level 4(:) 부연설명은 직전 태스크 요약에 붙인다.
        if task_text.startswith(":") and last_record is not None:
            continuation = _strip_bullet_prefix(task_text)
            if continuation:
                last_record["summary"] = f"{str(last_record.get('summary') or '').rstrip()}, {continuation}"
            continue

        if not _is_valid_status_text(status_cell) or not task_text:
            continue

        # 작성가이드 Level2(>)를 프로젝트 컨텍스트로 사용한다.
        if task_text.startswith(">"):
            level2_project = task_text.lstrip(">").strip()
            if level2_project and not _is_member_name(level2_project):
                current_project = level2_project

        if task_text.startswith(">"):
            project_name = _strip_bullet_prefix(task_text)
        elif current_project and current_project not in {"기타", "학습"}:
            project_name = current_project
        else:
            project_name = _strip_bullet_prefix(task_text) or "미분류 프로젝트"

        members = _extract_member_names(task_text, summary_cell, project_cell)
        member_name = ""
        if current_member:
            if known_members:
                preferred = [m for m in members if m in known_members]
                member_name = ", ".join(preferred) if preferred else current_member
            else:
                member_name = current_member
        else:
            member_name = ", ".join(members) if members else ""
        if not member_name:
            anomalies.append({"row": row_idx, "error": "담당자명을 식별하지 못했습니다."})
            continue
        center_name = current_center or inferred_center
        if not center_name:
            anomalies.append({"row": row_idx, "error": "센터명을 식별하지 못했습니다."})
            continue

        style_delayed = _row_has_red_style(
            list(row_cells),
            [col_status, col_summary, col_project],
        )
        extra_start, extra_end = _extract_date_range_from_text(task_text, row_text)
        prev_start_iso = _parse_short_date(prev_start_cell) or extra_start
        prev_end_iso = _parse_short_date(prev_end_cell) or extra_end
        next_start_iso = _parse_short_date(next_start_cell)
        next_end_iso = _parse_short_date(next_end_cell)
        is_delayed = status_cell == "지연" or style_delayed
        if not is_delayed and status_cell in ("진행", "진행중", "진행 중") and prev_end_iso:
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}", prev_end_iso):
                try:
                    is_delayed = datetime.strptime(prev_end_iso, "%Y-%m-%d").date() < date.today()
                except Exception:
                    pass

        records.append(
            {
                "__row_idx": row_idx,
                "project_name": project_name,
                "center_name": center_name,
                "member_name": member_name,
                "status": status_cell,
                "summary": _strip_bullet_prefix(task_text),
                "prev_start": prev_start_iso or None,
                "prev_end": prev_end_iso or None,
                "next_start": next_start_iso or None,
                "next_end": next_end_iso or None,
                "is_delayed": is_delayed,
                "has_issue_keyword": any(kw in task_text for kw in _ISSUE_KEYWORDS),
            }
        )
        last_record = records[-1]

    return records, anomalies


def _parse_excel_file(resolved_path: str) -> dict:
    """Excel 파일을 파싱하여 정규화된 레코드를 반환한다.

    아티팩트 해소나 state 저장 없이 순수하게 파일만 파싱한다.
    fanout.py의 병렬 파싱 워커에서 재사용된다.

    Args:
        resolved_path: 파싱할 Excel 파일의 절대 경로.

    Returns:
        성공 시 ``{"records": [...], "total_count": int, "anomalies": [...]}``,
        실패 시 ``{"error": "..."}``
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl이 설치되지 않았습니다. `uv add openpyxl` 후 재시도하세요."}

    path = Path(resolved_path).expanduser()
    if not path.exists() and resolved_path:
        # 절대경로 전달 실패 시 파일명 기준으로 uploads/다운로드 디렉토리를 보조 탐색한다.
        import os as _os
        basename = path.name
        fallback_candidates: list[Path] = []
        if basename:
            fallback_candidates.extend(
                sorted(Path("uploads").glob(f"*{basename}*"), key=lambda p: p.stat().st_mtime, reverse=True)
            )
            downloads_dir = Path(_os.getenv("EXCEL_FALLBACK_DIR", str(Path.home() / "Downloads")))
            if downloads_dir.exists():
                fallback_candidates.extend(
                    sorted(downloads_dir.glob(f"*{basename}*"), key=lambda p: p.stat().st_mtime, reverse=True)
                )
        if fallback_candidates:
            path = fallback_candidates[0]

    if not path.exists():
        return {"error": f"파일을 찾을 수 없습니다: {resolved_path}"}
    if path.suffix.lower() not in _SUPPORTED_EXCEL_EXTENSIONS:
        return {"error": f"지원하지 않는 파일 형식입니다: {path.suffix}. xlsx 파일을 사용하세요."}

    inferred_center = _infer_center_from_filename(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    if not inferred_center:
        inferred_center = _infer_center_from_sheet(ws)

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    col_map = _build_column_map(headers)
    required = ["project_name", "status", "summary"]
    missing = [f for f in required if f not in col_map]

    header_records: list[dict] = []
    header_anomalies: list[dict] = []
    if not missing:
        for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
            row_values = [c.value for c in row_cells]
            if all(v is None for v in row_values):
                continue
            row_data = dict(zip(headers, row_values))
            try:
                style_delayed = _row_has_red_style(
                    list(row_cells),
                    [
                        col_map.get("status", -1),
                        col_map.get("summary", -1),
                        col_map.get("project_name", -1),
                    ],
                )
                record = _parse_row(row_data, col_map, inferred_center, style_delayed=style_delayed)
                if record:
                    header_records.append(record)
                else:
                    header_anomalies.append(
                        {"row": row_idx, "error": "필수 필드(project/status/summary/center/member) 누락"}
                    )
            except Exception as exc:
                header_anomalies.append({"row": row_idx, "error": str(exc)})

    # 비정형(헤더 누락) 파일에서는 member 헤더 매핑 오탐이 커서 identity 보정을 비활성화한다.
    if missing:
        identity_map: dict[int, dict[str, str]] = {}
        identity_anomalies: list[dict] = []
    else:
        identity_map, identity_anomalies = _collect_header_identity_map(ws, headers, col_map)
    hierarchical_records, hierarchical_anomalies = _parse_hierarchical_rows(ws, inferred_center)

    # member는 헤더 기반 값을 우선 적용한다. center는 파일명 기준으로 고정한다.
    merged_hierarchical_records: list[dict] = []
    for record in hierarchical_records:
        row_idx = int(record.get("__row_idx") or 0)
        identity = identity_map.get(row_idx, {})
        if inferred_center:
            record["center_name"] = inferred_center
        if identity.get("member_name"):
            record["member_name"] = identity["member_name"]
        record.pop("__row_idx", None)
        merged_hierarchical_records.append(record)

    # 헤더명 매칭은 보조 신호로만 사용하고, 레코드 건수가 더 많은 경로를 채택한다.
    if len(merged_hierarchical_records) >= len(header_records):
        records = merged_hierarchical_records
        anomalies = hierarchical_anomalies + header_anomalies + identity_anomalies
    else:
        records = header_records
        anomalies = header_anomalies + hierarchical_anomalies + identity_anomalies

    if inferred_center:
        for record in records:
            record["center_name"] = inferred_center

    if not records:
        missing_note = f"헤더 매핑 누락: {missing}" if missing else "헤더 매핑은 가능했지만 레코드 변환 실패"
        return {
            "error": (
                "엑셀 구조를 원하는 보고 형식으로 변환하지 못했습니다. "
                f"{missing_note}. anomalies를 확인해 주세요."
            ),
            "anomalies": anomalies[:20],
        }

    result: dict[str, Any] = {
        "records": records,
        "total_count": len(records),
        "anomalies": anomalies,
    }
    try:
        validated = [TaskItem.model_validate(record).model_dump(mode="json") for record in records]
        result["records"] = validated
        result["total_count"] = len(validated)
    except ValidationError as exc:
        return {
            "error": f"파싱 결과 검증 실패(TaskItem): {_format_validation_error(exc)}",
            "anomalies": anomalies[:20],
        }

    return result


async def parse_and_analyze_tool(
    file_path: str = "",
    tool_context: ToolContext | None = None,
) -> dict:
    """Parse uploaded Excel and convert to normalized task records."""
    if tool_context is None:
        return {"error": "tool_context가 없어 파일을 처리할 수 없습니다."}

    resolved_path = file_path.strip()
    # LLM/tool 호출 시 경로가 따옴표 또는 file:// 형태로 들어오는 경우를 정리한다.
    if resolved_path.startswith(("\"", "'")) and resolved_path.endswith(("\"", "'")) and len(resolved_path) >= 2:
        resolved_path = resolved_path[1:-1].strip()
    if resolved_path.startswith("file://"):
        resolved_path = resolved_path[7:]
    resolved_path = resolved_path.replace("\u00a0", " ").strip()

    artifact_name: str | None = None
    if not resolved_path:
        resolved_path, artifact_name = await _resolve_excel_path_from_artifact(tool_context)
        if not resolved_path:
            return {
                "error": (
                    "분석할 파일이 없습니다. UI에서 엑셀 파일을 첨부한 뒤 다시 요청하거나 "
                    "`file_path` 인자를 전달하세요."
                )
            }

    result = _parse_excel_file(resolved_path)

    # artifact에서 생성된 임시 파일은 파싱 후 즉시 삭제한다.
    if artifact_name:
        try:
            Path(resolved_path).unlink(missing_ok=True)
        except Exception:
            pass

    if result.get("error"):
        return result

    if artifact_name:
        result["artifact_name"] = artifact_name

    tool_context.state["parsed_records"] = result
    return result

